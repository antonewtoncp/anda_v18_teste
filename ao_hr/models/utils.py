import pandas as pd
from io import BytesIO
import os
from datetime import datetime
import openpyxl
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
bio = BytesIO()
from openpyxl.styles import Border, Side

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
def payment_anda_partner(data_values, sheet_name):
    _TEMPLATE = {'NR': [], 'NOME': [], 'BANCO': [], 'NIB': [], 'VALOR': []}
    for data in data_values:
        _TEMPLATE['NR'].append(data['n'])
        _TEMPLATE['NOME'].append(data['name'])
        _TEMPLATE['BANCO'].append(data['bank'])
        _TEMPLATE['NIB'].append(data['nib'])
        _TEMPLATE['VALOR'].append(data['amount'])

    pd.set_option('max_colwidth', None)
    df = pd.DataFrame(data=_TEMPLATE)
    pd.set_option('max_colwidth', None)
    df.to_excel(os.path.expanduser('~' + '/payment_map_bank.xlsx'), sheet_name=sheet_name, index=False)


def generate_xls(data_values, sheet_name):
    template = {'NR ORD': [], 'NOME': [], 'CODIGO DEP': [], 'Nº INSS': [], 'POSICAO': [], 'SALARIO BASE': [],
                'DESCONTO DE FALTAS E ATRASOS': [], 'Subsídio Alimentação': [], 'Subsídio Transporte': [],
                'Subsídio de Férias': [], 'Abono de Familia': [], 'Subsídio de Natal': [],
                'TOTAL SUBSIDIOS': [], 'SALARIO ILIQUIDO': [], 'TAXA SS': [], 'REMUNERAÇÃO TRIBUTAVEL': [],
                'IRT': [], 'AVANCO SALARIO': [], 'EMPRESTIMO/LOAN': [], 'FATURAS DE SAÚDE': [], 'OUTROS DESCONTOS': [],
                'TOTAL DESCONTO': [], 'PERCENTAGEM %': [], 'SALARIO LIQUIDO': [], }
    for i in data_values:
        template['NR ORD'].append(i['n'])
        template['NOME'].append(i['name'])
        template['CODIGO DEP'].append(i['departament'])
        template['Nº INSS'].append(i['n_ss'])
        template['POSICAO'].append(i['job'])
        template['SALARIO BASE'].append(i['wage'])
        template['DESCONTO DE FALTAS E ATRASOS'].append(i['desc_faltas'])
        template['Subsídio Alimentação'].append(i['sub_ali'])
        template['Subsídio Transporte'].append(i['sub_trans'])
        template['Abono de Familia'].append(i['sub_fam'])
        template['Subsídio de Férias'].append(i['sub_ferias'])
        template['Subsídio de Natal'].append(i['sub_nat'])
        template['TOTAL SUBSIDIOS'].append(i['total_sub'])
        template['SALARIO ILIQUIDO'].append(i['gross_salary'])
        template['TAXA SS'].append(i['ss_tax'])
        template['REMUNERAÇÃO TRIBUTAVEL'].append(i['rem_tribut'])
        template['IRT'].append(i['irt'])
        template['AVANCO SALARIO'].append(i['salary_advance'])
        template['EMPRESTIMO/LOAN'].append(i['loan'])
        template['FATURAS DE SAÚDE'].append(i['fact_saude'])
        template['OUTROS DESCONTOS'].append(i['outros_desc'])
        template['TOTAL DESCONTO'].append(i['total_deductions'])
        template['PERCENTAGEM %'].append(i['percentage'])
        template['SALARIO LIQUIDO'].append(i['net_salary'])

    result = pd.DataFrame(data=template)
    result.to_excel('/home/aquima/salary_map.xlsx', sheet_name=sheet_name, index=False)


def read_xls():
    with open("/home/aquima/salary_map.xlsx", 'rb') as file:
        return file.read()


def read_payment_partner_xls():
    with open(os.path.expanduser('~' + '/payment_map_bank.xlsx'), 'rb') as file:
        return file.read()
    
def generate_bank_sheet(payment_map, company, sheet_name='Pagamento de Salário'):
    df = pd.DataFrame(payment_map["rows"])

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Formato de borda fina
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Largura das colunas (B até G ajustadas)
    ws.column_dimensions['A'].width = 2   # Margem vazia
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15  # Aqui vai o VALOR

    # Cabeçalho fixo
    ws['B2'] = company.name
    ws['B3'] = "NIB Ordenante"
    ws['C3'] = '-'
    ws['B4'] = "Moeda de Origem"
    ws['C4'] = company.currency_id.name
    ws['B5'] = "Data de Processamento (dia-mes-ano)"
    ws['C5'] = payment_map["info"]["create_date"]
    ws['B6'] = "Código de Operação"
    ws['C6'] = "-"
    ws['B7'] = "Referência do Ordenante"
    ws['C7'] = payment_map["info"]["reference"]

    # Alinhamento + borda no cabeçalho
    for row in range(2, 8):
        ws[f'C{row}'].alignment = Alignment(horizontal="right")
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border

    # Dados (começando da linha 9)
    start_row = 9
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for j, value in enumerate(row, start=2):  # Começa na coluna B até G
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border

    # Total
    total_row = start_row + len(df) + 1
    ws[f'B{total_row}'] = "Total"
    ws[f'B{total_row}'].border = thin_border

    total_cell = ws[f'G{total_row}']   # agora o total fica em G
    total_cell.value = df['valor'].sum()
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="center")
    total_cell.border = thin_border

    # Salvar na memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return base64.b64encode(output.read())
